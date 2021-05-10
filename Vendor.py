import openpyxl
import smptlib

pending_payment={}

def getDefaulters():
	
	global pending_payment
	excelObj=openpyxl.load_workbook('Vendor_payment.xlsx')  # load workbook which contains data of vendors and their payment history
 
	sheet=excelObj.active.title
	
	for row in range(2,excelObj[sheet].max_row+1):
		for column in range(2,excelObj[sheet].max_column+1):
		     
			if excelObj[sheet].cell(row=row,column=column).value==None: # check if vendor has not paid
				pending_payment.update({excelObj[sheet].cell(row=row,column=2).value:[]})	# get the email id of defaulters and store in dictionary

				

	# traverse the email ids of defaulters.
	for key in pending_payment.keys():
	     
		# get the payment pending months. 
		pending_payment[key]=[excelObj[sheet].cell(row=1,column=column).value for row in range(2,excelObj[sheet].max_row+1) for column in range(3,excelObj[sheet].max_column+1) if excelObj[sheet].cell(row=row,column=2).value == key and excelObj[sheet].cell(row=row,column=column).value==None]
		
	
	
	return(pending_payment)

# to check the defaulters and send them a reminder mail
def sendMail(defaulter):
	
	smtpObj=smtplib.SMTP('smtp.gmail.com',587)
	smtpObj.starttls()
	
	user_email=input('Enter your email')  # get a sender_mail
	password = input('Enter your password') #get a user password
	
	smtpObj.login(user_email,password)
	
	try:
		if smtpObj.login(user_email,password)[0]==235:  # To check if sender email and password are correct
		# send an automated reminder mail to all the defaulters.
			for key in defaulter.keys():

				smtpObj.sendmail(user_email,key,'Subject:Urgent Reminder!!\n Dear vendor,It is to remind you that you not behind your payments from the month of '+defaulter[key][0]+'\n Kindly pay your dues in order to continue further services. Thanks and Regards,Kapil')
		
			
	except smtplib.SMTPAuthenticationError:
		
		print("Please, check sender email and password and Try Again!!")
	
getDefaulters()
sendMail(getDefaulters())        			
